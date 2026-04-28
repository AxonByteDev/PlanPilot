"""Excel-Export aus der SQLite-Datenbank."""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

COLORS = {
    "sofort":       ("F8D7DA","721C24"),
    "kurzfristig":  ("FFF3CD","856404"),
    "mittelfristig":("D1ECF1","0C5460"),
    "langfristig":  ("D4EDDA","155724"),
}
LABELS = {
    "sofort":"🔴 Sofort","kurzfristig":"🟡 Kurzfristig",
    "mittelfristig":"🔵 Mittelfristig","langfristig":"🟢 Langfristig",
}
STATUS_COLORS = {
    "offen":   ("FFF3CD","856404"),
    "laufend": ("BEE5EB","0C5460"),
    "erledigt":("C3E6CB","155724"),
    "pausiert":("E2E3E5","383D41"),
}

def _fill(c): return PatternFill("solid", fgColor=c)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Segoe UI", italic=italic)
def _align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

def _cell(ws, ref, value="", bold=False, fg="000000", bg=None,
          h="left", v="top", size=10, italic=False, border=True):
    c = ws[ref]
    c.value = value
    c.font  = _font(bold, fg, size, italic)
    c.alignment = _align(h=h, v=v)
    if bg: c.fill = _fill(bg)
    if border: c.border = _border()


def build_excel(tasks, steps_by_task):
    wb = Workbook()

    # ── Übersicht ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Übersicht"
    for col, w in zip("ABCDEFGH", [5,28,16,14,50,22,14,12]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    _cell(ws,"A1","Projektplan – Rießelmann",bold=True,fg="FFFFFF",bg="1A3A5C",h="center",v="center",size=14,border=False)
    ws.row_dimensions[1].height = 28

    for col, h in zip("ABCDEFGH",["#","Aufgabe","Priorität","Aufwand","Schritte (Kurzform)","Abhängigkeiten","Zuständig","Status"]):
        _cell(ws,f"{col}2",h,bold=True,fg="FFFFFF",bg="1A3A5C",h="center",v="center")
    ws.row_dimensions[2].height = 18

    row = 3
    for i, t in enumerate(tasks):
        pbg, pfg = COLORS.get(t["priority"], ("FFFFFF","000000"))
        sbg, sfg = STATUS_COLORS.get(t["status"], ("FFFFFF","000000"))
        bg = "FFFFFF" if i % 2 == 0 else "F4F6F9"
        ws.row_dimensions[row].height = 70

        steps  = steps_by_task.get(t["id"], [])
        steps_text = "\n".join(f"{s['step_nr']}. {s['description']}" for s in steps[:6])
        if len(steps) > 6: steps_text += f"\n… (+{len(steps)-6} weitere)"

        _cell(ws,f"A{row}",t["nr"],bold=True,fg="888888",bg=bg,h="center",v="center")
        _cell(ws,f"B{row}",t["title"],bold=True,fg="000000",bg=bg,v="top")
        _cell(ws,f"C{row}",LABELS.get(t["priority"],t["priority"]),bold=True,fg=pfg,bg=pbg,h="center",v="center")
        _cell(ws,f"D{row}",t.get("aufwand",""),fg="1A3A5C",bg=bg,h="center",v="center")
        _cell(ws,f"E{row}",steps_text,fg="333333",bg=bg,v="top",size=9)
        _cell(ws,f"F{row}",t.get("dependencies",""),fg="C0392B",bg=bg,v="top",size=9,italic=True)
        _cell(ws,f"G{row}",t.get("assignee",""),bg=bg,v="center")
        _cell(ws,f"H{row}",t.get("status","offen").capitalize(),bold=True,fg=sfg,bg=sbg,h="center",v="center")
        row += 1

    ws.auto_filter.ref = f"A2:H{row-1}"
    ws.freeze_panes = "A3"

    # ── Detailplan ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detailplan")
    for col, w in zip("ABCDEFGHI",[5,26,16,55,16,16,14,10,12]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:I1")
    _cell(ws2,"A1","Detailplan – Einzelschritte",bold=True,fg="FFFFFF",bg="1A3A5C",h="center",v="center",size=14,border=False)
    ws2.row_dimensions[1].height = 28

    for col, h in zip("ABCDEFGHI",["#","Aufgabe","Priorität","Schritt","Geplant bis","Erledigt am","Zuständig","Aufwand","Status"]):
        _cell(ws2,f"{col}2",h,bold=True,fg="FFFFFF",bg="1A3A5C",h="center",v="center")
    ws2.row_dimensions[2].height = 18
    ws2.freeze_panes = "A3"

    row2 = 3
    for t in tasks:
        pbg, pfg = COLORS.get(t["priority"], ("FFFFFF","000000"))
        steps = steps_by_task.get(t["id"], [])

        ws2.merge_cells(f"A{row2}:I{row2}")
        _cell(ws2,f"A{row2}",f"#{t['nr']} – {t['title']}  |  {LABELS.get(t['priority'],'')}",
              bold=True,fg=pfg,bg=pbg,h="left",v="center",border=False)
        ws2.row_dimensions[row2].height = 16
        row2 += 1

        for j, s in enumerate(steps):
            bg = "FFFFFF" if j % 2 == 0 else "F4F6F9"
            sbg, sfg = STATUS_COLORS.get(s.get("status","offen"),("FFF3CD","856404"))
            ws2.row_dimensions[row2].height = 16
            _cell(ws2,f"A{row2}",t["nr"],fg="AAAAAA",bg=bg,h="center",v="center")
            _cell(ws2,f"B{row2}",t["title"],fg="666666",bg=bg,v="center",size=9,italic=True)
            _cell(ws2,f"C{row2}",LABELS.get(t["priority"],""),fg=pfg,bg=pbg,h="center",v="center",size=9)
            _cell(ws2,f"D{row2}",f"{s['step_nr']}. {s['description']}",fg="222222",bg=bg,v="center")
            _cell(ws2,f"E{row2}",s.get("geplant_bis",""),bg=bg,h="center",v="center")
            _cell(ws2,f"F{row2}",s.get("erledigt_am",""),bg=bg,h="center",v="center")
            _cell(ws2,f"G{row2}",s.get("assignee",""),bg=bg,v="center")
            _cell(ws2,f"H{row2}",s.get("aufwand",""),fg="1A3A5C",bg=bg,h="center",v="center")
            _cell(ws2,f"I{row2}",s.get("status","offen").capitalize(),bold=True,fg=sfg,bg=sbg,h="center",v="center")
            row2 += 1
        row2 += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
